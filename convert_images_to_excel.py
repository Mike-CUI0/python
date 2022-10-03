# 여러파일을 불러와서 처리하기
# 기본적인 엑셀로 변환 아이디어는 유투브 텍코딩이라는 분의 https://www.youtube.com/shorts/k0hjVTN68d0 를 참조했꼬
# 파일 여는 방법과 처리에 대해서는 https://www.youtube.com/watch?v=H71ts4XxWYU  Jie Jenn 을 참조 했다.

from configparser import Interpolation
import cv2  # opencv 이미지 프로세싱 관련 모듈
import openpyxl  # 엑셀관련 모듈
import datetime  # 시간관련하여 시간값을 가져오기 위한 함수. 나중에 파일명 저장할때 쓰려고 가져옴
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from matplotlib import colors  # 이미지 처리 및 데이터 처리 모듈 matplotlib 무지 유명

import tkinter as tk  # 파일다이어로그 관련해서 tkinter 를 약어 tk 로 지정.  파이썬2에서는 T가 대문자 였음
from tkinter import filedialog  # 파일 다이어로그만 가져온다
root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilenames()  # 유저가 불러온 파일명과 폴더명저장

for img_file in file_path:

    # 파일을 ㅔpath 가 걸려 있는 곳에다가 넣어야지만 읽어 들일수 있다. 파일이 없으면 에러발생
    # cv2 에서 이미지 읽는법이며 파일명읽고 1=컬러로 읽기 , 그레이스케일도 있으니 인터넷 찾아볼것
    img = cv2.imread(img_file, 1)
    # Mike 입장에서는 python_prg 가 python이 동작하는 패스가 되어서 그곳에다가 그림 파일 넣어야 하며 거기에다가 파일 저장함
    # cv2.resize 에 대한 매개변수 값도 구글링 해서 알아보면 됨
    img = cv2.resize(img, (160, 200), interpolation=cv2.INTER_AREA)

    wb = openpyxl.Workbook()  # 여기서 부터는 손쉽게 엑셀파일을 여는 openpyxl 에 관한 부분이니까 구글링에 다 있음
    sheet = wb.active
    sheet.title = "test"

    h, w, _ = img.shape

    for x in range(w):

        sheet.column_dimensions[get_column_letter(x+1)].width = 5
        for y in range(h):

            b, g, r = img[y, x]
            color_code = colors.to_hex([r/255, g/255, b/255])[1:]

            sheet.cell(row=y+1, column=x+1).fill = PatternFill(start_color=color_code,
                                                               end_color=color_code, fill_type="solid")

            sheet.row_dimensions[y].height = 30

    ctime = datetime.datetime.now()  # 시간값을 일단 저장
    str_ctime = str(ctime.year)+str(ctime.month)+str(ctime.day) + \
        "_"+str(ctime.hour) + str(ctime.minute) + str(ctime.second)
    # 저장한 시간값을 문자로 변환해서 저장할때 사용할 변수에 저장

    wb.save(img_file+"_"+str_ctime+".xlsx")  # wb.save openpyxl 기능

    print("처리 완료 되었습니다."+img_file+"_"+str_ctime+".xlsx로 저장하였습니다.")
    # 사용자가 처리 된 것을 알려 주기 위함
